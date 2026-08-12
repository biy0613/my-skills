# -*- coding: utf-8 -*-
"""포트폴리오기록.xlsx 생성 · 조회 · 기록 공용 라이브러리.

스킬 `stock-record` / `stock-price` 가 함께 쓴다.

워크북 구조
  대시보드  : KPI + 평가액 스냅샷 시계열 + 차트
  보유종목  : 종목 마스터 (수량·평단이 여기 소스 오브 트루스)
  시세기록  : 날짜×종목 주당단가 원장 (append-only) → 나머지가 여기서 파생
  설정      : 환율·비중 상한
  <종목명>  : 종목별 진입/보유점검/회수/사후추적 기록

종목 시트는 A열의 앵커 마커(#ENTRY, #REVIEW ...)로 섹션을 찾는다.
사용자가 행을 끼워 넣어도 기록 위치를 잃지 않게 하기 위함이다.
"""
from __future__ import annotations

import os
from datetime import date, datetime

from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# --------------------------------------------------------------------------
# 경로
# --------------------------------------------------------------------------
DEFAULT_WB_PATH = os.environ.get("STOCK_PORTFOLIO_XLSX") or os.path.join(
    os.path.expanduser("~"),
    "OneDrive", "바탕 화면", "클로드실습", "주식", "트레이딩 기록", "포트폴리오기록.xlsx",
)

# --------------------------------------------------------------------------
# 디자인 토큰
# --------------------------------------------------------------------------
INK = "1F2A44"        # 웜 네이비
ACCENT = "E8674C"     # 코랄
CREAM = "FBF7F0"
BAND = "EEF1F6"       # 섹션 밴드
LINE = "D6DBE4"
MUTED = "8A93A3"

F_TITLE = Font(name="맑은 고딕", size=16, bold=True, color=INK)
F_SECTION = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
F_HEAD = Font(name="맑은 고딕", size=9, bold=True, color="FFFFFF")
F_LABEL = Font(name="맑은 고딕", size=9, bold=True, color=INK)
F_BODY = Font(name="맑은 고딕", size=9, color="222222")
F_MUTED = Font(name="맑은 고딕", size=8, color=MUTED)
F_KPI_LABEL = Font(name="맑은 고딕", size=8, bold=True, color=MUTED)
F_KPI_VALUE = Font(name="맑은 고딕", size=13, bold=True, color=INK)
F_ANCHOR = Font(name="맑은 고딕", size=6, color="EDEDED")

FILL_HEAD = PatternFill("solid", fgColor=INK)
FILL_SECTION = PatternFill("solid", fgColor=INK)
FILL_BAND = PatternFill("solid", fgColor=BAND)
FILL_CREAM = PatternFill("solid", fgColor=CREAM)
FILL_INPUT = PatternFill("solid", fgColor="FFFFFF")
FILL_ACCENT = PatternFill("solid", fgColor="FDE8E2")   # 누적 손익 강조

THIN = Side(style="thin", color=LINE)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP_TOP = Alignment(wrap_text=True, vertical="top")
WRAP_CTR = Alignment(wrap_text=True, vertical="center", horizontal="center")
RIGHT = Alignment(horizontal="right", vertical="center")
LEFT_C = Alignment(horizontal="left", vertical="center")

# 숫자 포맷 — 한국 관행(상승 빨강 / 하락 파랑)
FMT_KRW = "#,##0"
FMT_PNL = '[Red]#,##0;[Blue]-#,##0;"-"'
FMT_PCT = '[Red]0.00%;[Blue]-0.00%;"-"'
FMT_PCT0 = "0.0%"
FMT_PRC = "#,##0.00"
FMT_DATE = "yyyy-mm-dd"

MAXR = 200            # 마스터/원장 수식이 훑는 최대 행

# --------------------------------------------------------------------------
# 시트별 스키마
# --------------------------------------------------------------------------
MASTER_COLS = [
    ("상태", 8), ("종목명", 14), ("티커", 10), ("시장", 9), ("통화", 7),
    ("수량", 10), ("평균단가", 12), ("투자원금(원)", 14),
    ("현재가", 12), ("시세기준일", 12), ("평가액(원)", 14),
    ("평가손익(원)", 14), ("수익률", 10), ("비중", 9), ("목표비중", 9),
    ("손절선", 11), ("익절목표", 11), ("손절까지", 10),
    ("진입일", 12), ("최근점검일", 12), ("기록시트", 14),
]
MASTER_HEAD_ROW = 2
MASTER_FIRST_ROW = 3

PRICE_COLS = [
    ("날짜", 12), ("종목명", 14), ("통화", 7), ("주당단가", 13),
    ("수량", 10), ("환율", 10), ("평가액(원)", 14), ("입력경로", 11),
]
PRICE_HEAD_ROW = 2
PRICE_FIRST_ROW = 3

SNAP_COLS = [
    ("기록일", 12), ("주식평가액", 15), ("보유현금", 14), ("총자산", 15),
    ("입출금(당회)", 13), ("순입금누계", 15), ("누적 손익", 15),
    ("누적수익률", 12), ("직전대비", 11), ("메모", 40),
]
SNAP_HEAD_ROW = 11
SNAP_FIRST_ROW = 12

# 종목 시트 — 진입 블록 항목
ENTRY_FIELDS = [
    "기록일",
    "하드게이트 통과 (O/6)",
    "미통과 항목",
    "규칙 이탈 여부",
    "핵심 논지 (한 문장)",
    "가정 1",
    "가정 2",
    "가정 3",
    "반대 논지 (2개 이상)",
    "앞으로의 기대감",
    "컨센서스와 내 견해 차이",
    "지금 가격이 이미 반영한 것",
    "펀더멘탈",
    "법률·규제 리스크",
    "수급",
    "매수 시 감정 상태",
    "목표 비중 / 최대 허용 손실",
    "분할·추가매수 규칙",
    "손절 조건 — 사실",
    "손절 조건 — 가격",
    "손절하지 않는 경우",
    "익절 조건 — 사실",
    "익절 조건 — 가격",
]

REVIEW_COLS = [
    ("점검일", 12), ("점검 요약 — 무엇이 바뀌었나", 70), ("가정1", 10),
    ("가정2", 10), ("가정3", 10), ("근거 변경 유무", 13), ("현가 매수 의중", 13),
    ("손절선 갱신", 12), ("현재 비중", 10), ("결론", 12),
]

EXIT_FIELDS = [
    "매도일",
    "매도 단가",
    "매도 수량",
    "실현 손익(원)",
    "실현 수익률",
    "보유 기간",
    "A 보류게이트 X 개수",
    "B 강제트리거 O 개수",
    "매도 유형",
    "유형 근거 — 요건이 성립하는가",
    "24h 쿨다운 했나",
    "계획대로였나",
    "이탈 내용",
]

HOLD_COLS = [
    ("날짜", 12), ("그래도 보유하기로 한 이유", 70), ("켜진 B 트리거", 22),
    ("구분", 14), ("재점검 기한", 12),
]

FOLLOW_COLS = [
    ("시점", 10), ("내 매도 판단은 옳았나", 70), ("날짜", 12),
    ("주가", 12), ("매도가 대비", 12),
]
FOLLOW_ROWS = ["1달 후", "2달 후", "반년 후"]

POST_FIELDS = [
    "과정/결과 4분면",
    "배운 것",
    "다음 종목에 적용할 규칙 변경",
]

# 드롭다운
DV_STATUS = '"보유,청산,관찰"'
DV_ASSUM = '"확인,미확인,반증"'
DV_YN = '"예,아니오"'
DV_DRIFT = '"같음,다름"'
DV_RESET = '"산다,안 산다"'
DV_REVIEW_CONC = '"유지,비중조절,회수검토,전량매도"'
DV_EXITTYPE = '"①논지훼손,②목표달성,③교체,④리스크관리,⑤현금필요,⑥타이밍"'
DV_PLAN = '"계획대로,규칙 이탈"'
DV_HOLDKIND = '"규칙 이탈,사전 정의된 예외"'
DV_QUAD = '"과정O·결과O,과정O·결과X,과정X·결과O(경고),과정X·결과X"'

STAGES = ("진입", "보유점검", "회수", "홀드", "사후추적")

# --------------------------------------------------------------------------
# 게이트 문항 — 셀 메모로 붙고, 스킬이 사용자에게 물을 때도 그대로 읽어준다
# --------------------------------------------------------------------------
ENTRY_GATES = [
    "실적 발표를 1회 이상 통과하며 관찰했는가 (최소 1개월)",
    "이 종목을 한 문장으로 설명할 수 있는가",
    "반대 논지를 2개 이상 적을 수 있는가",
    "손절 조건을 가격 + 사실 양쪽으로 숫자화했는가",
    "포지션 크기가 사전 규칙 내인가 (비중 상한 30% + 최대 허용 손실 30%)",
    "사는 이유가 급등·뉴스·손실 복구가 아닌가 (24시간 쿨다운)",
]

EXIT_GATES_A = [
    "매도 유형을 하나로 특정했는가",
    "진입 시 적어둔 손절/익절 조건 중 실제로 발동한 항목이 있는가",
    "매도 이유가 주가가 아니라 사실로 설명되는가",
    "이 결정을 어제도 똑같이 했을 것인가 (24h 쿨다운)",
    "오늘 안 갖고 있다면 이 가격에 사지 않을 것인가",
    "매도 후 2배 더 올라도 판단이 옳았다고 말할 근거가 있는가",
    "지금 감정이 공포·지루함·본전 심리가 아닌가",
]

EXIT_TRIGGERS_B = [
    "진입 시 적은 사실 기준(thesis break) 중 하나 이상 발생",
    "가정 3개 중 2개 이상 반증",
    '가격 손절선 이탈 + "손절하지 않는 경우"에 해당하지 않음',
    "경영진 신뢰 훼손 (회계 이슈, 공시 번복, 내부자 대량 매도, 갑작스러운 CFO 사임)",
    "논지 드리프트 — 새 이유로는 진입 게이트를 통과 못 함",
    "비중 상한 초과 상태로 3개월 이상 방치",
]

EXIT_TYPES = [
    ("①논지훼손", "사실 기준·가정이 실제로 반증됨. 가격 하락은 요건이 아님"),
    ("②목표달성", "익절 사실 기준 도달 또는 목표가 도달 + 남은 논지 없음"),
    ("③교체", "새 후보가 진입 게이트 통과 + 세금·비용 차감 후에도 우위"),
    ("④리스크관리", "논지는 유효하나 비중 상한 초과·유동성 문제. 부분 매도가 원칙"),
    ("⑤현금필요", "투자 판단과 무관한 실생활 자금 수요"),
    ("⑥타이밍", "논지는 그대로인데 가격·시점 판단으로 매도. **재매수 의도가 있는 매도**. "
              "고르면 재매수 조건(가격·시점)을 숫자로 함께 적어야 하고, "
              "사후추적에서 실제 재매수 여부를 확인한다"),
]


def _numbered(items) -> str:
    return "\n".join(f"{i}. {t}" for i, t in enumerate(items, 1))


CELL_NOTES = {
    "하드게이트 통과 (O/6)": (
        "진입 하드게이트 6 — 하나라도 X면 매수 보류\n\n"
        + _numbered(ENTRY_GATES)
        + "\n\nX가 나오면 X라고 적는다. 그래도 진행하면 '규칙 이탈 여부'에 사유를 적는다."
    ),
    "A 보류게이트 X 개수": (
        "A 보류게이트 7 — 하나라도 X면 매도 보류\n\n"
        + _numbered(EXIT_GATES_A)
        + "\n\n6번은 익절 전용 장치다. 손절은 규칙으로 잡히지만\n"
          "너무 빨리 파는 익절은 규칙에 안 걸린다(처분효과)."
    ),
    "B 강제트리거 O 개수": (
        "B 강제트리거 6 — 하나라도 O면 매도가 원칙\n\n"
        + _numbered(EXIT_TRIGGERS_B)
        + "\n\nB가 켜졌는데 A가 X면 B가 우선한다.\n"
          "A는 충동 매도를 막는 장치지 강제 매도를 무르는 장치가 아니다.\n"
          "B가 켜졌는데 팔지 않았다면 반드시 [③-E 홀드 기록]을 남긴다."
    ),
    "매도 유형": (
        "매도 유형 6 — 하나만 고른다\n\n"
        + "\n".join(f"{k} : {v}" for k, v in EXIT_TYPES)
        + "\n\n①~⑤는 '논지에 무슨 일이 생겼나'를 묻는다.\n"
          "⑥만 논지가 그대로인 채 '잠깐 나갔다 들어온다'는 매도다.\n\n"
          "두 개 이상 해당한다고 느껴지면 진짜 이유를 다른 이름으로\n"
          "부르고 있다는 신호다.\n"
          "흔한 오분류 — 손실 회피를 ③으로 포장 / 전량 매도해놓고 ④라고 적음 /\n"
          "재매수 조건 없이 ⑥이라 적기(그건 그냥 매도다)"
    ),
    "24h 쿨다운 했나": (
        "매도 결정을 하루 재우고 실행했는가.\n"
        "급락·급등 당일 판단을 차단하는 장치다 (A 보류게이트 4번)."
    ),
}


# --------------------------------------------------------------------------
# 저수준 헬퍼
# --------------------------------------------------------------------------
def _anchor(ws, row: int, marker: str) -> None:
    """A열에 섹션 앵커를 심는다 (거의 안 보이는 회색 6pt)."""
    c = ws.cell(row=row, column=1, value=marker)
    c.font = F_ANCHOR


def _section(ws, row: int, title: str, last_col: int = 10) -> None:
    for col in range(2, last_col + 1):
        c = ws.cell(row=row, column=col)
        c.fill = FILL_SECTION
    c = ws.cell(row=row, column=2, value=title)
    c.font = F_SECTION
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[row].height = 22


def _table_header(ws, row: int, cols, start_col: int = 2) -> None:
    for i, (name, width) in enumerate(cols):
        col = start_col + i
        c = ws.cell(row=row, column=col, value=name)
        c.font = F_HEAD
        c.fill = FILL_HEAD
        c.alignment = WRAP_CTR
        c.border = BOX
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[row].height = 26


def _label_value_block(ws, start_row: int, fields, note: str | None = None) -> int:
    """B=라벨 / C=입력칸 세로 블록. 마지막 행 번호를 돌려준다."""
    r = start_row
    for f in fields:
        lab = ws.cell(row=r, column=2, value=f)
        lab.font = F_LABEL
        lab.fill = FILL_BAND
        lab.alignment = WRAP_TOP
        lab.border = BOX
        if f in CELL_NOTES:
            lab.comment = Comment(CELL_NOTES[f], "체크포인트", width=420, height=230)
            lab.font = Font(name="맑은 고딕", size=9, bold=True, color=ACCENT)
        val = ws.cell(row=r, column=3)
        val.font = F_BODY
        val.alignment = WRAP_TOP
        val.border = BOX
        val.fill = FILL_INPUT
        ws.row_dimensions[r].height = 30
        r += 1
    if note:
        c = ws.cell(row=r, column=2, value=note)
        c.font = F_MUTED
        r += 1
    return r


def _blank_rows(ws, header_row: int, cols, n: int, start_col: int = 2) -> None:
    for i in range(n):
        r = header_row + 1 + i
        for j in range(len(cols)):
            c = ws.cell(row=r, column=start_col + j)
            c.border = BOX
            c.font = F_BODY
            c.alignment = WRAP_TOP
        ws.row_dimensions[r].height = 28


def _add_dv(ws, formula: str, cell_range: str) -> None:
    dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(cell_range)


def _safe_sheet_name(name: str) -> str:
    bad = set('[]:*?/\\')
    cleaned = "".join(ch for ch in name if ch not in bad).strip()
    return cleaned[:31] or "종목"


def _today() -> str:
    return date.today().isoformat()


def _norm_date(v):
    """문자열/날짜 무엇이 오든 date 로."""
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip().replace("/", "-").replace(".", "-")
    for fmt in ("%Y-%m-%d", "%y-%m-%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s[:19], fmt).date()
        except ValueError:
            continue
    return None


# --------------------------------------------------------------------------
# 시트 빌더
# --------------------------------------------------------------------------
def build_settings(wb) -> None:
    ws = wb.create_sheet("설정")
    ws.sheet_properties.tabColor = MUTED
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 60

    ws["B2"] = "설정"
    ws["B2"].font = F_TITLE

    rows = [
        ("기준 통화", "KRW", "모든 평가액은 원화로 환산해 집계한다"),
        ("USD/KRW 환율", 1380, "stock-price 실행 시 함께 갱신한다"),
        ("환율 기준일", None, ""),
        ("단일 종목 비중 상한", 0.30, "진입 게이트 5번 판정 기준"),
        ("단일 팩터 비중 상한", 0.50, "분기 포트폴리오 점검용"),
        ("최대 허용 손실 (1종목)", 0.30, "종목당 −30%. 진입 게이트 5번 판정 기준 (고정값)"),
    ]
    r = 4
    for lab, val, memo in rows:
        ws.cell(row=r, column=2, value=lab).font = F_LABEL
        ws.cell(row=r, column=2).fill = FILL_BAND
        ws.cell(row=r, column=2).border = BOX
        c = ws.cell(row=r, column=3, value=val)
        c.font = F_BODY
        c.border = BOX
        c.alignment = LEFT_C
        ws.cell(row=r, column=4, value=memo).font = F_MUTED
        r += 1

    ws["C5"].number_format = FMT_KRW
    ws["C6"].number_format = FMT_DATE
    for cell in ("C7", "C8", "C9"):
        ws[cell].number_format = FMT_PCT0

    wb.defined_names.add(_defined_name("환율", "설정", "$C$5"))
    wb.defined_names.add(_defined_name("비중상한", "설정", "$C$7"))

    ws["B12"] = "이 파일을 고치는 방법"
    ws["B12"].font = F_LABEL
    for i, line in enumerate([
        "· 수량·평균단가는 [보유종목] 시트에서만 고친다. 나머지 시트는 전부 여기서 파생된 수식이다.",
        "· 주당단가는 [시세기록]에 쌓기만 하면 된다. 가장 마지막 행이 현재가로 잡힌다.",
        "· 종목별 시트의 A열 회색 글자(#ENTRY 등)는 스킬이 기록 위치를 찾는 앵커다. 지우지 말 것.",
        "· 행을 끼워 넣거나 지워도 앵커만 살아 있으면 자동 기록은 계속 동작한다.",
    ]):
        ws.cell(row=13 + i, column=2, value=line).font = F_MUTED


def _defined_name(name, sheet, ref):
    from openpyxl.workbook.defined_name import DefinedName
    return DefinedName(name, attr_text=f"{sheet}!{ref}")


def build_master(wb) -> None:
    ws = wb.create_sheet("보유종목")
    ws.sheet_properties.tabColor = INK
    ws.column_dimensions["A"].width = 3
    ws["B1"] = "보유종목 마스터 — 수량·평균단가는 여기서만 고친다"
    ws["B1"].font = F_TITLE
    ws.row_dimensions[1].height = 26

    _table_header(ws, MASTER_HEAD_ROW, MASTER_COLS)
    ws.freeze_panes = "D3"
    last_col = get_column_letter(1 + len(MASTER_COLS))
    ws.auto_filter.ref = f"B{MASTER_HEAD_ROW}:{last_col}{MASTER_HEAD_ROW}"

    for r in range(MASTER_FIRST_ROW, MASTER_FIRST_ROW + 40):
        _write_master_formulas(ws, r)

    _add_dv(ws, DV_STATUS, f"B{MASTER_FIRST_ROW}:B{MASTER_FIRST_ROW+39}")


def _write_master_formulas(ws, r: int) -> None:
    """마스터 한 행의 서식과 파생 수식."""
    for i in range(len(MASTER_COLS)):
        c = ws.cell(row=r, column=2 + i)
        c.border = BOX
        c.font = F_BODY
        c.alignment = LEFT_C
    ws.row_dimensions[r].height = 20

    # 마스터 열 매핑: B상태 C종목명 D티커 E시장 F통화 G수량 H평균단가 I투자원금
    #                J현재가 K시세기준일 L평가액 M평가손익 N수익률 O비중 P목표비중
    #                Q손절선 R익절목표 S손절까지 T진입일 U최근점검일 V기록시트
    # 시세기록 열 매핑: B날짜 C종목명 D통화 E주당단가 F수량 G환율 H평가액 I입력경로
    name = f"$C{r}"
    fx = "설정!$C$5"
    ws[f"I{r}"] = (f'=IF(OR($G{r}="",$H{r}=""),"",'
                   f'IF($F{r}="USD",$G{r}*$H{r}*{fx},$G{r}*$H{r}))')
    ws[f"J{r}"] = (f'=IFERROR(LOOKUP(2,1/(시세기록!$C$3:$C${MAXR}={name}),'
                   f"시세기록!$E$3:$E${MAXR}),\"\")")
    ws[f"K{r}"] = (f'=IFERROR(LOOKUP(2,1/(시세기록!$C$3:$C${MAXR}={name}),'
                   f"시세기록!$B$3:$B${MAXR}),\"\")")
    ws[f"L{r}"] = (f'=IF(OR($G{r}="",$J{r}=""),"",'
                   f'IF($F{r}="USD",$G{r}*$J{r}*{fx},$G{r}*$J{r}))')
    ws[f"M{r}"] = f'=IF(OR($L{r}="",$I{r}=""),"",$L{r}-$I{r})'
    ws[f"N{r}"] = f'=IFERROR($M{r}/$I{r},"")'
    # 비중 분모는 총자산(주식평가액 + 보유현금). 현금을 빼면 비중이 과대평가된다.
    ws[f"O{r}"] = f'=IFERROR($L{r}/대시보드!$E$5,"")'
    ws[f"S{r}"] = f'=IFERROR($Q{r}/$J{r}-1,"")'
    ws[f"V{r}"] = f'=IF($C{r}="","",$C{r})'

    ws[f"G{r}"].number_format = "#,##0.####"
    for col in ("H", "J", "Q", "R"):
        ws[f"{col}{r}"].number_format = FMT_PRC
    for col in ("I", "L"):
        ws[f"{col}{r}"].number_format = FMT_KRW
    ws[f"M{r}"].number_format = FMT_PNL
    for col in ("N", "S"):
        ws[f"{col}{r}"].number_format = FMT_PCT
    for col in ("O", "P"):
        ws[f"{col}{r}"].number_format = FMT_PCT0
    for col in ("K", "T", "U"):
        ws[f"{col}{r}"].number_format = FMT_DATE
    for col in ("I", "J", "K", "L", "M", "N", "O", "S", "V"):
        ws[f"{col}{r}"].fill = FILL_CREAM


def build_prices(wb) -> None:
    ws = wb.create_sheet("시세기록")
    ws.sheet_properties.tabColor = ACCENT
    ws.column_dimensions["A"].width = 3
    ws["B1"] = "시세 원장 — 아래로 계속 쌓기만 한다 (stock-price 스킬이 append)"
    ws["B1"].font = F_TITLE
    ws.row_dimensions[1].height = 26

    _table_header(ws, PRICE_HEAD_ROW, PRICE_COLS)
    ws.freeze_panes = "B3"
    last_col = get_column_letter(1 + len(PRICE_COLS))
    ws.auto_filter.ref = f"B{PRICE_HEAD_ROW}:{last_col}{PRICE_HEAD_ROW}"

    for r in range(PRICE_FIRST_ROW, MAXR + 1):
        _write_price_row_format(ws, r)


def _write_price_row_format(ws, r: int) -> None:
    # 열: B날짜 C종목명 D통화 E주당단가 F수량 G환율 H평가액 I입력경로
    for i in range(len(PRICE_COLS)):
        c = ws.cell(row=r, column=2 + i)
        c.border = BOX
        c.font = F_BODY
        c.alignment = LEFT_C
    ws[f"B{r}"].number_format = FMT_DATE
    ws[f"E{r}"].number_format = FMT_PRC
    ws[f"F{r}"].number_format = "#,##0.####"
    ws[f"G{r}"].number_format = FMT_KRW
    ws[f"H{r}"] = (f'=IF(OR($E{r}="",$F{r}=""),"",'
                   f'IF($D{r}="USD",$E{r}*$F{r}*$G{r},$E{r}*$F{r}))')
    ws[f"H{r}"].number_format = FMT_KRW
    ws[f"H{r}"].fill = FILL_CREAM
    ws.row_dimensions[r].height = 18


def dashboard_kpis():
    """대시보드 KPI — (자산 밴드, 손익 밴드) 두 줄로 돌려준다.

    자산 밴드 (라벨 4행 / 값 5행): B주식평가액 C보유현금 D현금비중 E총자산
                                  F보유종목수 G최근기록일
    손익 밴드 (라벨 6행 / 값 7행): B순입금누계 C누적손익 D누적수익률
                                  E실현손익 F투자원금 G보유종목평가손익 H보유종목수익률

    보유종목 시트의 '비중' 이 **E5(총자산)** 를 분모로 쓴다. 순서를 바꾸면 거기도 고칠 것.

    두 손익을 굳이 나눠 두는 이유: '누적 손익'은 계좌 전체(입금액 대비) 성적이고
    '보유종목 평가손익'은 지금 들고 있는 포지션의 성적이다. 실현 손절이 있으면 둘이
    크게 벌어지는데, 섞어 보면 계좌가 깨진 걸 포지션 수익률로 가려 버리게 된다.
    """
    mlast = MASTER_FIRST_ROW + 39
    lookup_last = lambda col: (f'=IFERROR(LOOKUP(2,1/($B${SNAP_FIRST_ROW}:$B${MAXR}<>""),'
                               f'${col}${SNAP_FIRST_ROW}:${col}${MAXR}),0)')
    asset = [
        ("주식 평가액", f'=SUMIF(보유종목!$B$3:$B${mlast},"보유",보유종목!$L$3:$L${mlast})', FMT_KRW),
        ("보유 현금", lookup_last("D"), FMT_KRW),
        ("현금 비중", '=IFERROR($C$5/$E$5,"")', FMT_PCT0),
        ("총자산", "=$B$5+$C$5", FMT_KRW),
        ("보유 종목 수", f'=COUNTIF(보유종목!$B$3:$B${mlast},"보유")', "0"),
        ("최근 기록일", lookup_last("B").replace(",0)", ',"")'), FMT_DATE),
    ]
    pnl = [
        ("순입금 누계", lookup_last("G"), FMT_KRW),
        ("누적 손익", "=$E$5-$B$7", FMT_PNL),
        ("누적 수익률", '=IFERROR($C$7/$B$7,"")', FMT_PCT),
        ("실현 손익", "=$C$7-$G$7", FMT_PNL),
        ("투자원금(보유분)", f'=SUMIF(보유종목!$B$3:$B${mlast},"보유",보유종목!$I$3:$I${mlast})', FMT_KRW),
        ("보유종목 평가손익", "=$B$5-$F$7", FMT_PNL),
        ("보유종목 수익률", '=IFERROR($G$7/$F$7,"")', FMT_PCT),
    ]
    return asset, pnl


def write_kpi_block(ws) -> None:
    """KPI 두 밴드와 스냅샷 제목을 (재)작성한다. 전부 수식이라 덮어써도 안전하다."""
    asset, pnl = dashboard_kpis()

    def band(items, label_row, value_row, highlight=None):
        for i, (label, formula, fmt) in enumerate(items):
            col = 2 + i
            lc = ws.cell(row=label_row, column=col, value=label)
            lc.font = F_KPI_LABEL
            lc.fill = FILL_CREAM
            lc.alignment = Alignment(horizontal="center", vertical="center")
            lc.border = BOX
            vc = ws.cell(row=value_row, column=col, value=formula)
            vc.font = F_KPI_VALUE
            vc.fill = FILL_ACCENT if label == highlight else FILL_CREAM
            vc.number_format = fmt
            vc.alignment = Alignment(horizontal="center", vertical="center")
            vc.border = BOX
            ws.column_dimensions[get_column_letter(col)].width = 17
        ws.row_dimensions[label_row].height = 18
        ws.row_dimensions[value_row].height = 30
        # 이전 레이아웃이 더 길었을 수 있다. 밴드 오른쪽 잔재를 지운다.
        for col in range(2 + len(items), 16):
            for row in (label_row, value_row):
                c = ws.cell(row=row, column=col)
                c.value = None
                c.fill = PatternFill(fill_type=None)
                c.border = Border()

    band(asset, 4, 5)
    band(pnl, 6, 7, highlight="누적 손익")

    # 날짜는 13pt 굵게면 폭이 모자라 ####### 이 된다
    ws.cell(row=5, column=1 + len(asset)).font = Font(
        name="맑은 고딕", size=11, bold=True, color=INK)
    ws.cell(row=4, column=1, value="자산").font = F_MUTED
    ws.cell(row=6, column=1, value="손익").font = F_MUTED

    ws.cell(row=8, column=2, value=None)
    ws.cell(row=9, column=2, value="■ 평가액 스냅샷").font = Font(
        name="맑은 고딕", size=12, bold=True, color=INK)
    ws.cell(row=10, column=2, value=(
        "주 2회 stock-price 스킬로 append. 주식평가액은 [시세기록]의 같은 날짜 합계로 자동 계산된다. "
        "보유현금과 입출금만 직접 넣으면 된다.")).font = F_MUTED


def build_dashboard(wb) -> None:
    ws = wb.create_sheet("대시보드", 0)
    ws.sheet_properties.tabColor = ACCENT
    ws.column_dimensions["A"].width = 3
    ws.sheet_view.showGridLines = False

    ws["B2"] = "포트폴리오 기록"
    ws["B2"].font = Font(name="맑은 고딕", size=20, bold=True, color=INK)
    ws["B3"] = "수량·평단은 [보유종목], 주당단가는 [시세기록]에 넣는다. 이 시트는 전부 수식이다."
    ws["B3"].font = F_MUTED
    ws.row_dimensions[2].height = 30

    write_kpi_block(ws)   # KPI 두 밴드 + 스냅샷 제목(9·10행)까지 여기서 쓴다

    _table_header(ws, SNAP_HEAD_ROW, SNAP_COLS)
    for r in range(SNAP_FIRST_ROW, MAXR + 1):
        _write_snapshot_row_format(ws, r)
    ws.freeze_panes = f"B{SNAP_FIRST_ROW}"

    chart = LineChart()
    chart.title = "총자산 추이"
    chart.height = 8.5
    chart.width = 20
    chart.style = 2
    data = Reference(ws, min_col=5, min_row=SNAP_HEAD_ROW, max_row=SNAP_FIRST_ROW + 60)
    cats = Reference(ws, min_col=2, min_row=SNAP_FIRST_ROW, max_row=SNAP_FIRST_ROW + 60)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.y_axis.numFmt = "#,##0"
    ws.add_chart(chart, "M11")


def _write_snapshot_row_format(ws, r: int) -> None:
    # 열: B기록일 C주식평가액 D보유현금 E총자산 F입출금 G순입금누계 H평가손익 I누적수익률 J직전대비 K메모
    for i in range(len(SNAP_COLS)):
        c = ws.cell(row=r, column=2 + i)
        c.border = BOX
        c.font = F_BODY
        c.alignment = LEFT_C
    ws[f"B{r}"].number_format = FMT_DATE
    for col in ("C", "D", "E", "F", "G"):
        ws[f"{col}{r}"].number_format = FMT_KRW
    ws[f"H{r}"].number_format = FMT_PNL
    for col in ("I", "J"):
        ws[f"{col}{r}"].number_format = FMT_PCT
    ws[f"K{r}"].alignment = WRAP_TOP
    ws.row_dimensions[r].height = 18


def snapshot_row_formulas(r: int) -> dict:
    """스냅샷 한 행에 들어갈 파생 수식 (append 시 스크립트가 씀)."""
    return {
        "C": f'=IF($B{r}="","",IFERROR(SUMIFS(시세기록!$H$3:$H${MAXR},'
             f'시세기록!$B$3:$B${MAXR},$B{r}),0))',
        "E": f'=IF($B{r}="","",N($C{r})+N($D{r}))',
        "G": f'=IF($B{r}="","",SUM($F${SNAP_FIRST_ROW}:$F{r}))',
        "H": f'=IF($B{r}="","",$E{r}-$G{r})',
        "I": f'=IFERROR($H{r}/$G{r},"")',
        "J": (f'=IFERROR($E{r}/$E{r-1}-1,"")' if r > SNAP_FIRST_ROW else '=""'),
    }


# --------------------------------------------------------------------------
# 종목 시트
# --------------------------------------------------------------------------
def build_stock_sheet(wb, name: str):
    """종목 기록 시트를 만든다. 이미 있으면 그대로 돌려준다."""
    sname = _safe_sheet_name(name)
    if sname in wb.sheetnames:
        return wb[sname]

    ws = wb.create_sheet(sname)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 72
    for col in "DEFGHIJ":
        ws.column_dimensions[col].width = 13
    ws.column_dimensions["K"].width = 13

    ws["B2"] = name
    ws["B2"].font = F_TITLE
    ws["C2"] = "← 이 셀의 종목명으로 [보유종목] 마스터를 찾는다. 마스터의 종목명과 정확히 같아야 한다."
    ws["C2"].font = F_MUTED
    ws.row_dimensions[2].height = 26

    mlast = MASTER_FIRST_ROW + 39

    def midx(col_letter: str) -> str:
        return (f'=IFERROR(INDEX(보유종목!${col_letter}$3:${col_letter}${mlast},'
                f'MATCH($B$2,보유종목!$C$3:$C${mlast},0)),"")')

    r = 4
    _anchor(ws, r, "#SUMMARY")
    _section(ws, r, "요약 — [보유종목] 마스터에서 자동")
    r += 1
    summary = [
        ("상태", midx("B"), None),
        ("수량", midx("G"), "#,##0.####"),
        ("평균단가", midx("H"), FMT_PRC),
        ("현재가", midx("J"), FMT_PRC),
        ("평가액(원)", midx("L"), FMT_KRW),
        ("평가손익(원)", midx("M"), FMT_PNL),
        ("수익률", midx("N"), FMT_PCT),
        ("비중", midx("O"), FMT_PCT0),
        ("진입일", midx("T"), FMT_DATE),
        ("최근 점검일", midx("U"), FMT_DATE),
    ]
    for lab, formula, fmt in summary:
        lc = ws.cell(row=r, column=2, value=lab)
        lc.font = F_LABEL
        lc.fill = FILL_BAND
        lc.border = BOX
        vc = ws.cell(row=r, column=3, value=formula)
        vc.font = F_BODY
        vc.border = BOX
        vc.fill = FILL_CREAM
        vc.alignment = LEFT_C
        if fmt:
            vc.number_format = fmt
        ws.row_dimensions[r].height = 18
        r += 1

    r += 1
    _anchor(ws, r, "#ENTRY")
    _section(ws, r, "① 진입 기록 — 매수 주문 전에 작성. 체결 후 수정 금지(추가만)")
    r += 1
    r = _label_value_block(ws, r, ENTRY_FIELDS)

    r += 1
    _anchor(ws, r, "#REVIEW")
    _section(ws, r, "② 보유 점검 — 분기 1회 + 이벤트 발생 시 (아래로 누적)")
    r += 1
    hdr = r
    _table_header(ws, hdr, REVIEW_COLS)
    _blank_rows(ws, hdr, REVIEW_COLS, 12)
    _add_dv(ws, DV_ASSUM, f"D{hdr+1}:F{hdr+12}")
    _add_dv(ws, DV_DRIFT, f"G{hdr+1}:G{hdr+12}")
    _add_dv(ws, DV_RESET, f"H{hdr+1}:H{hdr+12}")
    _add_dv(ws, DV_REVIEW_CONC, f"K{hdr+1}:K{hdr+12}")
    for rr in range(hdr + 1, hdr + 13):
        ws[f"B{rr}"].number_format = FMT_DATE
        ws[f"J{rr}"].number_format = FMT_PCT0
    r = hdr + 13

    r += 1
    _anchor(ws, r, "#EXIT")
    _section(ws, r, "③ 회수 — A 보류게이트 7 / B 강제트리거 6 통과 후 작성")
    r += 1
    exit_start = r
    r = _label_value_block(ws, r, EXIT_FIELDS)

    def erow(label: str) -> int:
        return exit_start + EXIT_FIELDS.index(label)

    ws[f"C{erow('매도일')}"].number_format = FMT_DATE
    ws[f"C{erow('매도 단가')}"].number_format = FMT_PRC
    ws[f"C{erow('실현 손익(원)')}"].number_format = FMT_PNL
    ws[f"C{erow('실현 수익률')}"].number_format = FMT_PCT
    _add_dv(ws, DV_EXITTYPE, f"C{erow('매도 유형')}")
    _add_dv(ws, DV_YN, f"C{erow('24h 쿨다운 했나')}")
    _add_dv(ws, DV_PLAN, f"C{erow('계획대로였나')}")

    r += 1
    _anchor(ws, r, "#HOLD")
    _section(ws, r, "③-E 홀드 기록 — B 트리거가 켜졌는데 팔지 않은 경우 (반드시 기록)")
    r += 1
    hdr = r
    _table_header(ws, hdr, HOLD_COLS)
    _blank_rows(ws, hdr, HOLD_COLS, 6)
    _add_dv(ws, DV_HOLDKIND, f"E{hdr+1}:E{hdr+6}")
    for rr in range(hdr + 1, hdr + 7):
        ws[f"B{rr}"].number_format = FMT_DATE
        ws[f"F{rr}"].number_format = FMT_DATE
    r = hdr + 7

    r += 1
    _anchor(ws, r, "#FOLLOW")
    _section(ws, r, "④ 사후 추적 — 매도 후에도 닫지 않는다")
    r += 1
    hdr = r
    _table_header(ws, hdr, FOLLOW_COLS)
    _blank_rows(ws, hdr, FOLLOW_COLS, 3)
    for i, lab in enumerate(FOLLOW_ROWS):
        ws.cell(row=hdr + 1 + i, column=2, value=lab).font = F_LABEL
        ws[f"D{hdr+1+i}"].number_format = FMT_DATE
        ws[f"E{hdr+1+i}"].number_format = FMT_PRC
        ws[f"F{hdr+1+i}"].number_format = FMT_PCT
    r = hdr + 4

    r += 1
    _anchor(ws, r, "#POST")
    _section(ws, r, "④-B 사후 평가")
    r += 1
    post_start = r
    r = _label_value_block(ws, r, POST_FIELDS)
    _add_dv(ws, DV_QUAD, f"C{post_start}")

    _anchor(ws, r + 1, "#END")
    return ws


# --------------------------------------------------------------------------
# 워크북 생성
# --------------------------------------------------------------------------
def build_workbook(path: str, holdings: list[dict] | None = None) -> str:
    wb = Workbook()
    wb.remove(wb.active)

    build_dashboard(wb)
    build_master(wb)
    build_prices(wb)
    build_settings(wb)

    holdings = holdings or []
    ws = wb["보유종목"]
    for i, h in enumerate(holdings):
        r = MASTER_FIRST_ROW + i
        ws[f"B{r}"] = h.get("상태", "보유")
        ws[f"C{r}"] = h["종목명"]
        ws[f"D{r}"] = h.get("티커", "")
        ws[f"E{r}"] = h.get("시장", "")
        ws[f"F{r}"] = h.get("통화", "KRW")
        ws[f"G{r}"] = h.get("수량")
        ws[f"H{r}"] = h.get("평균단가")
        build_stock_sheet(wb, h["종목명"])

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    return path


# --------------------------------------------------------------------------
# 조회 / 기록 API
# --------------------------------------------------------------------------
def open_wb(path: str = None):
    return load_workbook(path or DEFAULT_WB_PATH)


def read_holdings(path: str = None, only_active: bool = True) -> list[dict]:
    """마스터에서 종목 목록을 읽는다 (수식이 아닌 원시 입력값만)."""
    wb = load_workbook(path or DEFAULT_WB_PATH, data_only=False)
    ws = wb["보유종목"]
    out = []
    for r in range(MASTER_FIRST_ROW, MASTER_FIRST_ROW + 40):
        name = ws[f"C{r}"].value
        if not name:
            continue
        status = ws[f"B{r}"].value or "보유"
        if only_active and status != "보유":
            continue
        out.append({
            "row": r,
            "상태": status,
            "종목명": name,
            "티커": ws[f"D{r}"].value,
            "시장": ws[f"E{r}"].value,
            "통화": ws[f"F{r}"].value or "KRW",
            "수량": ws[f"G{r}"].value,
            "평균단가": ws[f"H{r}"].value,
        })
    return out


def get_setting(wb, key_row: int):
    return wb["설정"].cell(row=key_row, column=3).value


def find_anchor(ws, marker: str) -> int | None:
    for r in range(1, ws.max_row + 2):
        if ws.cell(row=r, column=1).value == marker:
            return r
    return None


def first_empty_row(ws, header_row: int, key_col: int = 2, limit: int = 400) -> int:
    r = header_row + 1
    while r < header_row + limit:
        if ws.cell(row=r, column=key_col).value in (None, ""):
            return r
        r += 1
    return r


def _ensure_row_style(ws, row: int, cols, start_col: int = 2) -> None:
    for j in range(len(cols)):
        c = ws.cell(row=row, column=start_col + j)
        if c.border is None or c.border.left.style is None:
            c.border = BOX
        c.font = F_BODY
        c.alignment = WRAP_TOP


def write_entry(ws, data: dict) -> list[str]:
    """진입 블록: 라벨 매칭으로 C열에 채운다."""
    start = find_anchor(ws, "#ENTRY")
    if start is None:
        raise RuntimeError("#ENTRY 앵커를 찾지 못했다")
    written = []
    r = start + 1
    for _ in range(len(ENTRY_FIELDS) + 6):
        label = ws.cell(row=r, column=2).value
        if label in data and data[label] not in (None, ""):
            val = data[label]
            if label == "기록일":
                val = _norm_date(val) or val
            ws.cell(row=r, column=3, value=val)
            written.append(label)
        r += 1
        if ws.cell(row=r, column=1).value and str(ws.cell(row=r, column=1).value).startswith("#"):
            break
    return written


def write_exit(ws, data: dict) -> list[str]:
    start = find_anchor(ws, "#EXIT")
    if start is None:
        raise RuntimeError("#EXIT 앵커를 찾지 못했다")
    written = []
    r = start + 1
    for _ in range(len(EXIT_FIELDS) + 6):
        label = ws.cell(row=r, column=2).value
        if label in data and data[label] not in (None, ""):
            val = data[label]
            if label == "매도일":
                val = _norm_date(val) or val
            ws.cell(row=r, column=3, value=val)
            written.append(label)
        r += 1
        if ws.cell(row=r, column=1).value and str(ws.cell(row=r, column=1).value).startswith("#"):
            break
    return written


def append_review(ws, data: dict) -> int:
    hdr = find_anchor(ws, "#REVIEW")
    if hdr is None:
        raise RuntimeError("#REVIEW 앵커를 찾지 못했다")
    hdr += 1
    row = first_empty_row(ws, hdr)
    _ensure_row_style(ws, row, REVIEW_COLS)
    order = ["점검일", "점검 요약", "가정1", "가정2", "가정3",
             "근거 변경 유무", "현가 매수 의중", "손절선 갱신", "현재 비중", "결론"]
    for j, key in enumerate(order):
        v = data.get(key)
        if key == "점검일":
            v = _norm_date(v) or v
        ws.cell(row=row, column=2 + j, value=v)
    ws.cell(row=row, column=2).number_format = FMT_DATE
    ws.cell(row=row, column=10).number_format = FMT_PCT0
    ws.cell(row=row, column=3).alignment = WRAP_TOP
    ws.row_dimensions[row].height = 40
    return row


def append_hold(ws, data: dict) -> int:
    hdr = find_anchor(ws, "#HOLD")
    if hdr is None:
        raise RuntimeError("#HOLD 앵커를 찾지 못했다")
    hdr += 1
    row = first_empty_row(ws, hdr)
    _ensure_row_style(ws, row, HOLD_COLS)
    order = ["날짜", "보유 이유", "켜진 B 트리거", "구분", "재점검 기한"]
    for j, key in enumerate(order):
        v = data.get(key)
        if key in ("날짜", "재점검 기한"):
            v = _norm_date(v) or v
        ws.cell(row=row, column=2 + j, value=v)
    ws.cell(row=row, column=2).number_format = FMT_DATE
    ws.cell(row=row, column=6).number_format = FMT_DATE
    ws.row_dimensions[row].height = 40
    return row


def append_follow(ws, data: dict) -> int:
    """사후추적: 시점(D+90 등) 행을 찾아 채운다. 없으면 새 행."""
    hdr = find_anchor(ws, "#FOLLOW")
    if hdr is None:
        raise RuntimeError("#FOLLOW 앵커를 찾지 못했다")
    hdr += 1
    target = str(data.get("시점", "")).strip()
    row = None
    for r in range(hdr + 1, hdr + 10):
        v = ws.cell(row=r, column=2).value
        if v is None:
            row = row or r
            break
        if str(v).strip() == target:
            row = r
            break
    if row is None:
        row = first_empty_row(ws, hdr)
    _ensure_row_style(ws, row, FOLLOW_COLS)
    ws.cell(row=row, column=2, value=target or ws.cell(row=row, column=2).value)
    ws.cell(row=row, column=3, value=data.get("판단 평가"))
    ws.cell(row=row, column=4, value=_norm_date(data.get("날짜")) or data.get("날짜"))
    ws.cell(row=row, column=5, value=data.get("주가"))
    ws.cell(row=row, column=6, value=data.get("매도가 대비"))
    ws.cell(row=row, column=4).number_format = FMT_DATE
    ws.cell(row=row, column=5).number_format = FMT_PRC
    ws.cell(row=row, column=6).number_format = FMT_PCT
    ws.cell(row=row, column=3).alignment = WRAP_TOP
    ws.row_dimensions[row].height = 40

    post = find_anchor(ws, "#POST")
    if post is not None:
        pr = post + 1
        for key in POST_FIELDS:
            label = ws.cell(row=pr, column=2).value
            if label in data and data[label] not in (None, ""):
                ws.cell(row=pr, column=3, value=data[label])
            pr += 1
    return row


# --------------------------------------------------------------------------
# 입력 템플릿 — "엑셀의 어느 행에 무엇을 적어야 하는가"
# --------------------------------------------------------------------------
FIELD_HINTS = {
    "기록일": "yyyy-mm-dd",
    "하드게이트 통과 (O/6)": "예: 6/6",
    "미통과 항목": "X가 난 번호와 이유. 없으면 '없음'",
    "규칙 이탈 여부": "없으면 '없음'",
    "핵심 논지 (한 문장)": "한 문장. 두 문장이면 아직 정리가 안 된 것",
    "가정 1": "논지를 지탱하는 가정",
    "가정 2": "논지를 지탱하는 가정",
    "가정 3": "논지를 지탱하는 가정",
    "반대 논지 (2개 이상)": "파는 쪽 논리를 가장 설득력 있게. 2개 이상",
    "앞으로의 기대감": "이벤트 / 성장 / 재평가",
    "컨센서스와 내 견해 차이": "비면 초과수익의 근거가 없다",
    "지금 가격이 이미 반영한 것": "현재 밸류에이션이 암시하는 성장률·마진",
    "펀더멘탈": "매출·이익 구조, 현금흐름, 부채, 희석",
    "법률·규제 리스크": "진행 사안, 최악 손실 규모, 일정. 없으면 '없음'",
    "수급": "외인·기관 지분율, 공매도 잔고, 내부자 매매",
    "매수 시 감정 상태": "평온 / 조급 / 복구심리 / FOMO",
    "목표 비중 / 최대 허용 손실": "총자산 대비 % + 손실 상한",
    "분할·추가매수 규칙": "지금 정하지 않으면 하락 중에 반드시 물타기로 간다",
    "손절 조건 — 사실": "무엇이 확인되면 논지가 죽나 (가격 아님)",
    "손절 조건 — 가격": "숫자",
    "손절하지 않는 경우": "예: 시장 전체 조정이고 위 사실 기준 미발생",
    "익절 조건 — 사실": "논지가 완성되는 사건",
    "익절 조건 — 가격": "목표가 + 근거",
    # 회수
    "매도 유형": "①논지훼손 / ②목표달성 / ③교체 / ④리스크관리 / ⑤현금필요 — 하나만",
    "유형 근거 — 요건이 성립하는가": "그 유형의 요건을 실제로 만족하는지",
    "24h 쿨다운 했나": "예 / 아니오",
    "계획대로였나": "계획대로 / 규칙 이탈",
    "A 보류게이트 X 개수": "7문항 중 X 개수",
    "B 강제트리거 O 개수": "6문항 중 O 개수",
    # 보유점검
    "가정1": "확인 / 미확인 / 반증",
    "가정2": "확인 / 미확인 / 반증",
    "가정3": "확인 / 미확인 / 반증",
    "근거 변경 유무": "같음 / 다름",
    "현가 매수 의중": "산다 / 안 산다",
    "손절선 갱신": "강화(위로)만 허용. 완화 금지",
    "결론": "유지 / 비중조절 / 회수검토 / 전량매도",
    # 홀드
    "구분": "규칙 이탈 / 사전 정의된 예외",
    "켜진 B 트리거": "번호",
    "재점검 기한": "최대 1개월",
    # 사후추적
    "시점": "1달 후 / 2달 후 / 반년 후",
    "과정/결과 4분면": "과정O·결과O / 과정O·결과X / 과정X·결과O(경고) / 과정X·결과X",
}

# 사용자에게 물을 문장. 라벨만 던지면 무엇을 적어야 할지 모른다.
FIELD_QUESTIONS = {
    # 진입 — C21~C39 (C17~C20은 게이트 답변에서 자동 생성)
    "핵심 논지 (한 문장)": "이 종목을 사는 이유를 **한 문장**으로 말하면?",
    "가정 1": "그 논지가 성립하려면 무엇이 참이어야 합니까? (가정 1)",
    "가정 2": "두 번째로 참이어야 하는 것은? (가정 2)",
    "가정 3": "세 번째로 참이어야 하는 것은? (가정 3)",
    "반대 논지 (2개 이상)": "지금 이 주식을 파는 사람은 무엇을 알고 있을까요? 가장 설득력 있는 반대 논지 2개 이상",
    "앞으로의 기대감": "앞으로 이 종목에서 무엇을 기대합니까? (이벤트 / 성장 / 재평가)",
    "컨센서스와 내 견해 차이": "시장이 믿고 있는 것과, 내가 다르게 보는 것은?",
    "지금 가격이 이미 반영한 것": "현재 주가는 이미 무엇을 반영하고 있습니까? (지금 밸류에이션이 암시하는 성장률·마진)",
    "펀더멘탈": "무엇을 팔아서 얼마를 법니까? (매출·이익 구조, 현금흐름, 부채, 희석)",
    "법률·규제 리스크": "진행 중인 소송·규제 이슈가 있습니까? 없으면 '없음'",
    "수급": "외국인·기관·개인이 이 주식을 어떻게 보고 있습니까? (지분율 추세, 공매도 잔고, 내부자 매매)",
    "매수 시 감정 상태": "살 때 심리 상태는? (평온 / 조급 / 복구심리 / FOMO)",
    "목표 비중 / 최대 허용 손실": "총자산 대비 몇 %까지 담을 계획이고, 이 종목에서 감당할 최대 손실은?",
    "분할·추가매수 규칙": "추가 매수는 어떤 조건에서 몇 회까지? (지금 안 정하면 하락 중에 물타기로 간다)",
    "손절 조건 — 사실": "**무슨 사실이 확인되면** 이 논지가 죽은 것입니까? (가격 아님)",
    "손절 조건 — 가격": "손절 가격은 얼마입니까?",
    "손절하지 않는 경우": "가격이 손절선에 닿아도 팔지 않을 경우는 언제입니까?",
    "익절 조건 — 사실": "무슨 일이 일어나면 논지가 완성된 것입니까?",
    "익절 조건 — 가격": "익절 가격과 그 근거는?",
    # 회수
    "매도일": "매도일은?",
    "매도 단가": "매도 단가는?",
    "매도 수량": "매도 수량은?",
    "실현 손익(원)": "실현 손익(원)은?",
    "실현 수익률": "실현 수익률은?",
    "보유 기간": "보유 기간은?",
    "매도 유형": ("매도 유형 ①~⑥ 중 하나는? (①논지훼손 ②목표달성 ③교체 ④리스크관리 "
               "⑤현금필요 ⑥타이밍) — ⑥이면 재매수 조건(가격·시점)도 함께"),
    "유형 근거 — 요건이 성립하는가": "그 유형의 요건이 실제로 성립합니까? 근거는?",
    "24h 쿨다운 했나": "매도 결정을 하루 재웠습니까? (예 / 아니오)",
    "계획대로였나": "계획대로였습니까, 규칙 이탈이었습니까?",
    "이탈 내용": "이탈이면 무엇을 어떻게 벗어났습니까?",
    # 보유점검
    "점검일": "점검일은?",
    "점검 요약 — 무엇이 바뀌었나": "진입 이후 무엇이 바뀌었습니까?",
    "가정1": "진입 시 가정 1의 상태는? (확인 / 미확인 / 반증)",
    "가정2": "가정 2의 상태는? (확인 / 미확인 / 반증)",
    "가정3": "가정 3의 상태는? (확인 / 미확인 / 반증)",
    "근거 변경 유무": "지금 들고 있는 이유가 살 때 이유와 같습니까? (같음 / 다름)",
    "현가 매수 의중": "오늘 안 갖고 있다면 이 가격에 살 것입니까? (산다 / 안 산다)",
    "손절선 갱신": "손절선을 바꿉니까? (강화만 허용, 완화 금지)",
    "현재 비중": "현재 비중은?",
    "결론": "결론은? (유지 / 비중조절 / 회수검토 / 전량매도)",
    # 홀드
    "날짜": "날짜는?",
    "보유 이유": "B 트리거가 켜졌는데도 계속 보유하기로 한 이유는?",
    "켜진 B 트리거": "켜진 B 트리거 번호는?",
    "구분": "규칙 이탈입니까, 사전 정의된 예외입니까?",
    "재점검 기한": "언제 다시 볼 것입니까? (최대 1개월)",
    # 사후추적
    "시점": "어느 시점입니까? (1달 후 / 2달 후 / 반년 후)",
    "주가": "그 시점 주가는?",
    "매도가 대비": "매도가 대비 몇 %입니까?",
    "내 매도 판단은 옳았나": "그때 매도 판단은 옳았습니까?",
    "과정/결과 4분면": "과정과 결과를 4분면으로 분류하면? (과정O·결과O / 과정O·결과X / 과정X·결과O(경고) / 과정X·결과X)",
    "배운 것": "배운 것은? (1~2줄)",
    "다음 종목에 적용할 규칙 변경": "다음 종목에 적용할 규칙 변경이 있습니까?",
}

# 게이트 답변에서 자동 생성되므로 사용자에게 묻지 않는 칸
AUTO_FIELDS = {
    "진입": ["기록일", "하드게이트 통과 (O/6)", "미통과 항목", "규칙 이탈 여부"],
    "회수": ["A 보류게이트 X 개수", "B 강제트리거 O 개수"],
    "보유점검": [],
    "홀드": [],
    "사후추적": [],
}

_STAGE_BLOCK = {"진입": ("#ENTRY", ENTRY_FIELDS), "회수": ("#EXIT", EXIT_FIELDS)}
_STAGE_TABLE = {"보유점검": ("#REVIEW", REVIEW_COLS),
                "홀드": ("#HOLD", HOLD_COLS),
                "사후추적": ("#FOLLOW", FOLLOW_COLS)}


def block_status(ws, stage: str, row: int | None = None) -> list[dict]:
    """해당 단계에서 채워야 할 자리를 엑셀 좌표와 함께 돌려준다.

    반환: [{"cell": "C33", "row": 33, "필드": "수급", "값": None, "힌트": "..."}]

    표 단계(보유점검·홀드·사후추적)는 기본적으로 **다음에 기록될 빈 행**을 본다.
    방금 기록한 행의 충실도를 보려면 `row` 로 그 행을 직접 지정한다 —
    지정하지 않으면 이미 쓴 행이 아니라 그 아래 빈 행을 검사해 전부 비었다고 보고한다.
    """
    out = []
    if stage in _STAGE_BLOCK:
        marker, fields = _STAGE_BLOCK[stage]
        start = find_anchor(ws, marker)
        if start is None:
            raise RuntimeError(f"{marker} 앵커를 찾지 못했다")
        for i, f in enumerate(fields):
            r = start + 1 + i
            out.append({"cell": f"C{r}", "row": r, "필드": f,
                        "값": ws.cell(row=r, column=3).value,
                        "힌트": FIELD_HINTS.get(f, "")})
        return out

    marker, cols = _STAGE_TABLE[stage]
    hdr = find_anchor(ws, marker)
    if hdr is None:
        raise RuntimeError(f"{marker} 앵커를 찾지 못했다")
    hdr += 1
    if row is not None:
        rows = [row]
    elif stage == "사후추적":
        rows = [hdr + 1 + i for i in range(len(FOLLOW_ROWS))]
    else:
        rows = [first_empty_row(ws, hdr)]
    for r in rows:
        for j, (name, _w) in enumerate(cols):
            col = 2 + j
            out.append({"cell": f"{get_column_letter(col)}{r}", "row": r, "필드": name,
                        "값": ws.cell(row=r, column=col).value,
                        "힌트": FIELD_HINTS.get(name, "")})
    return out


def render_template(ws, stage: str, stock: str) -> str:
    """채팅에 그대로 띄울 입력 폼. 빈 자리를 먼저, 이미 적힌 자리를 뒤에 보여준다."""
    st = block_status(ws, stage)
    empty = [x for x in st if x["값"] in (None, "")]
    filled = [x for x in st if x["값"] not in (None, "")]
    w = max((len(x["필드"]) for x in st), default=10)

    lines = [f"■ {stock} · {stage}  —  시트 [{ws.title}]",
             f"   빈 자리 {len(empty)} / 전체 {len(st)}", ""]
    if empty:
        lines.append(f"── 채워야 할 자리 {len(empty)}개 ──")
        for x in empty:
            hint = f"   ← {x['힌트']}" if x["힌트"] else ""
            lines.append(f"  {x['cell']:>5}  {x['필드']:<{w}} : {hint}")
    else:
        lines.append("── 빈 자리 없음 ──")
    if filled:
        lines += ["", f"── 이미 적힌 자리 {len(filled)}개 (고치려면 그대로 다시 주세요) ──"]
        for x in filled:
            v = str(x["값"]).replace("\n", " ")
            lines.append(f"  {x['cell']:>5}  {x['필드']:<{w}} : {v[:60]}{'…' if len(v) > 60 else ''}")
    return "\n".join(lines)


def render_questions(ws, stage: str, stock: str) -> str:
    """사용자가 바로 답할 수 있는 질문 목록. 자동 생성 칸은 빼고 번호를 매긴다."""
    st = block_status(ws, stage)
    auto = set(AUTO_FIELDS.get(stage, []))
    ask = [x for x in st if x["필드"] not in auto]
    cells = [x["cell"] for x in ask]
    span = f"{cells[0]}~{cells[-1]}" if cells else "-"

    lines = [f"■ {stock} · {stage}  —  답변할 항목 {len(ask)}개  (엑셀 [{ws.title}] {span})", ""]
    for i, x in enumerate(ask, 1):
        q = FIELD_QUESTIONS.get(x["필드"], x["힌트"] or x["필드"])
        lines.append(f"{i:2}. [{x['cell']}] {x['필드']}")
        lines.append(f"    {q}")
        cur = x["값"]
        if cur not in (None, ""):
            v = str(cur).replace("\n", " ")
            lines.append(f"    (현재: {v[:70]}{'…' if len(v) > 70 else ''})")
        lines.append("    →")
        lines.append("")
    if auto:
        lines.append(f"※ {', '.join(sorted(auto))} 은 게이트 답변에서 자동으로 채워집니다.")
    return "\n".join(lines)


def sync_alert_levels(wb, name: str, **levels) -> dict:
    """[알림설정] 시트의 가격 임계치를 갱신한다.

    종목 시트(서술)와 알림설정(자동 판정)이 따로 놀면 알림이 옛 값으로 돈다.
    실제로 2026-08-10 갱신값이 알림설정에 반영되지 않아 이틀간 틀린 손절선으로
    판정했다. 기록할 때 같이 쓰도록 여기로 묶는다.
    """
    if "알림설정" not in wb.sheetnames:
        return {}
    ws = wb["알림설정"]
    col = {"손절선": 7, "익절1": 8, "익절2": 9, "추가매수": 10}
    changed = {}
    for r in range(5, 40):
        n, watch = ws.cell(r, 2).value, ws.cell(r, 3).value
        if n != name or str(watch).upper() not in ("O", "X"):
            continue
        for k, v in levels.items():
            if v in (None, "") or k not in col:
                continue
            old = ws.cell(r, col[k]).value
            if old != v:
                ws.cell(r, col[k], value=v)
                changed[k] = (old, v)
        return changed
    return {}


def touch_master(ws_master, name: str, **fields) -> None:
    """마스터의 해당 종목 행 갱신 (진입일·최근점검일·상태·손절선 등)."""
    col_of = {
        "상태": "B", "티커": "D", "시장": "E", "통화": "F", "수량": "G",
        "평균단가": "H", "목표비중": "P", "손절선": "Q", "익절목표": "R",
        "진입일": "T", "최근점검일": "U",
    }
    for r in range(MASTER_FIRST_ROW, MASTER_FIRST_ROW + 40):
        if ws_master[f"C{r}"].value == name:
            for k, v in fields.items():
                if v in (None, "") or k not in col_of:
                    continue
                if k in ("진입일", "최근점검일"):
                    v = _norm_date(v) or v
                ws_master[f"{col_of[k]}{r}"] = v
            return
    # 없으면 새 행
    for r in range(MASTER_FIRST_ROW, MASTER_FIRST_ROW + 40):
        if not ws_master[f"C{r}"].value:
            ws_master[f"B{r}"] = fields.get("상태", "보유")
            ws_master[f"C{r}"] = name
            for k, v in fields.items():
                if v in (None, "") or k not in col_of:
                    continue
                if k in ("진입일", "최근점검일"):
                    v = _norm_date(v) or v
                ws_master[f"{col_of[k]}{r}"] = v
            return
    raise RuntimeError("마스터에 빈 행이 없다 — 행을 늘려야 한다")
